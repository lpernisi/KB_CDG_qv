# -*- coding: utf-8 -*-
"""
Genera il REPORT INVENTARIO DI BILANCIO (Excel) confrontando, riga per riga:
  - il VALORE del report Mago "Magazzino a valori" (giacenza alla data x prezzo bilancio)
  - il VALORE calcolato da NOI = giacenza del report x costo EFFICACE nostro (kodice.vw_costo_eff)
    con taglio valorizzazione APRILE 2026. Per i KIT il costo = distinta esplosa sui costi efficaci.

Costo efficace (vw_costo_eff): 1) ricalcolo wap_ricalc Aprile (split puro/oneri),
2) ripiego risalita su ultimo WAPCost>0 storico. Articoli senza nessuna fonte -> COSTO MANCANTE.

Uso:  python -m src.genera_report_inventario
Input report: arg 1 oppure il default qui sotto.
"""
import sys, os
from collections import defaultdict
from sqlalchemy import text
from src.config_loader import carica_config
from src import db

REPORT_TXT = sys.argv[1] if len(sys.argv) > 1 else \
    r"C:\Users\mago.admin\Downloads\Magazzino a valori 2026-06-04 SENZA WAP maggio.txt"
OUT_XLSX = r"C:\Users\mago.admin\Downloads\Inventario bilancio - confronto report vs nostro calcolo.xlsx"


def itnum(s):
    s = (s or "").strip()
    if s.count(",") == 1 and s.rfind(",") > s.rfind("."):
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def carica_report(path):
    righe = []
    with open(path, encoding="utf-8-sig") as fh:
        fh.readline()  # header
        for line in fh:
            p = line.rstrip("\n").split(";")
            if len(p) < 9:
                continue
            righe.append({
                "item": p[0].strip(), "descr": p[1].strip(), "giac": itnum(p[2]),
                "tipoV": p[3].strip(), "prezzo_rep": itnum(p[4]), "val_rep": itnum(p[6]),
                "cat": p[8].strip(),
            })
    return righe


def main():
    cfg = carica_config()
    eng = db._engine(cfg, cfg["database"]["dwh"])
    with eng.begin() as c:
        # costo efficace per articolo
        costo = {}
        for r in c.execute(text("SELECT Item,CostoEff,PuroUnit,OneriUnit,Fonte FROM kodice.vw_costo_eff")):
            costo[r[0]] = {"cost": float(r[1] or 0), "puro": float(r[2] or 0),
                           "oneri": float(r[3] or 0), "fonte": r[4]}
        # distinta per esplosione kit
        figli = defaultdict(list)
        for r in c.execute(text("SELECT LTRIM(RTRIM(BOM)),LTRIM(RTRIM(Component)),CAST(Qty AS float) FROM kodice.vw_distinta")):
            figli[r[0]].append((r[1], float(r[2] or 0)))
    kit_roots = set(figli.keys())

    # esplosione kit con costi efficaci (memoizzata, guardia anti-ciclo)
    memo = {}

    def costo_kit(bom, vis=None):
        if bom in memo:
            return memo[bom]
        vis = vis or set()
        if bom in vis:
            return (0.0, 0.0, True)  # ciclo
        vis = vis | {bom}
        tot_puro = tot_oneri = 0.0
        incompleto = False
        for comp, q in figli[bom]:
            if comp in figli:  # sotto-kit
                p, o, inc = costo_kit(comp, vis)
                tot_puro += p * q
                tot_oneri += o * q
                incompleto = incompleto or inc
            else:
                ce = costo.get(comp)
                if ce is None:
                    incompleto = True
                else:
                    tot_puro += ce["puro"] * q
                    tot_oneri += ce["oneri"] * q
        memo[bom] = (tot_puro, tot_oneri, incompleto)
        return memo[bom]

    righe = carica_report(REPORT_TXT)
    out = []
    for r in righe:
        it = r["item"]
        if it in kit_roots:
            puro, oneri, inc = costo_kit(it)
            ucost = puro + oneri
            fonte = "KIT_INCOMPLETO" if inc else "KIT"
            upuro, uoneri = puro, oneri
        else:
            ce = costo.get(it)
            if ce is not None and ce["cost"]:
                ucost, upuro, uoneri, fonte = ce["cost"], ce["puro"], ce["oneri"], ce["fonte"]
            elif r["prezzo_rep"]:  # ripiego: prezzo di bilancio del report Mago (medio non ricostruibile in SQL)
                ucost = upuro = r["prezzo_rep"]
                uoneri = 0.0
                fonte = "MEDIO_REPORT"
            else:
                ucost = upuro = uoneri = 0.0
                fonte = "MANCANTE"
        nostro_val = r["giac"] * ucost
        delta = nostro_val - r["val_rep"]
        deltap = (delta / r["val_rep"] * 100.0) if r["val_rep"] else None
        # anomalia prezzo report: abbiamo un costo INDIPENDENTE (non preso dal report) molto diverso
        # dal prezzo Mago -> il prezzo report e' sospetto (es. bidet a 2,49 vs nostro 60,74)
        sospetto = ""
        if fonte not in ("MEDIO_REPORT", "MANCANTE") and ucost > 0 and r["prezzo_rep"] > 0:
            if r["prezzo_rep"] < 0.5 * ucost or r["prezzo_rep"] > 2.0 * ucost:
                sospetto = "SI"
        out.append({
            "Articolo": it, "Descrizione": r["descr"], "Categoria": r["cat"], "TipoV": r["tipoV"],
            "Giacenza": r["giac"], "Prezzo report": r["prezzo_rep"], "Valore report": r["val_rep"],
            "Nostro costo unit.": ucost, "  di cui puro": upuro, "  di cui oneri": uoneri,
            "Fonte costo": fonte, "Nostro valore": nostro_val,
            "Delta (nostro-report)": delta, "Delta %": deltap,
            "Prezzo report sospetto": sospetto,
        })

    scrivi_excel(out)


def scrivi_excel(out):
    import pandas as pd
    df = pd.DataFrame(out)
    tot_rep = df["Valore report"].sum()
    tot_nos = df["Nostro valore"].sum()

    # riepilogo per fonte
    g = df.groupby("Fonte costo").agg(
        n=("Articolo", "count"),
        valore_report=("Valore report", "sum"),
        nostro_valore=("Nostro valore", "sum"),
    ).reset_index()
    g["delta"] = g["nostro_valore"] - g["valore_report"]

    # riepilogo per tipo valorizzazione
    gt = df.groupby("TipoV").agg(
        n=("Articolo", "count"),
        valore_report=("Valore report", "sum"),
        nostro_valore=("Nostro valore", "sum"),
    ).reset_index()
    gt["delta"] = gt["nostro_valore"] - gt["valore_report"]

    n_sosp = int((df["Prezzo report sospetto"] == "SI").sum())
    riepilogo = pd.DataFrame({
        "Voce": ["Totale VALORE REPORT (bilancio Mago)", "Totale NOSTRO VALORE (costo efficace x giac. report)",
                 "DELTA (nostro - report)", "Delta %", "Articoli totali",
                 "Articoli con costo MANCANTE", "Valore non coperto (MANCANTE)",
                 "Prezzi report SOSPETTI (da verificare)"],
        "Valore": [round(tot_rep, 2), round(tot_nos, 2), round(tot_nos - tot_rep, 2),
                   round((tot_nos - tot_rep) / tot_rep * 100, 3) if tot_rep else None,
                   len(df), int((df["Fonte costo"] == "MANCANTE").sum()),
                   round(df.loc[df["Fonte costo"] == "MANCANTE", "Valore report"].sum(), 2),
                   n_sosp],
    })

    # ordina il dettaglio per delta assoluto (le righe piu' divergenti in cima)
    df_ord = df.reindex(df["Delta (nostro-report)"].abs().sort_values(ascending=False).index)

    # foglio dedicato ai prezzi report sospetti, ordinati per impatto a valore
    sosp = df[df["Prezzo report sospetto"] == "SI"].copy()
    sosp["Scarto valore"] = sosp["Nostro valore"] - sosp["Valore report"]
    sosp = sosp.reindex(sosp["Scarto valore"].abs().sort_values(ascending=False).index)

    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as xw:
        riepilogo.to_excel(xw, sheet_name="Riepilogo", index=False)
        g.to_excel(xw, sheet_name="Riepilogo", index=False, startrow=len(riepilogo) + 3)
        gt.to_excel(xw, sheet_name="Riepilogo", index=False, startrow=len(riepilogo) + 3 + len(g) + 3)
        df_ord.to_excel(xw, sheet_name="Dettaglio", index=False)
        sosp.to_excel(xw, sheet_name="Prezzi report sospetti", index=False)

    # formattazione larghezze + numeri
    from openpyxl import load_workbook
    wb = load_workbook(OUT_XLSX)
    ws = wb["Dettaglio"]
    larg = {"A": 22, "B": 42, "C": 20, "D": 6, "E": 11, "F": 13, "G": 15,
            "H": 16, "I": 13, "J": 13, "K": 16, "L": 15, "M": 18, "N": 9, "O": 12}
    for ws in (wb["Dettaglio"], wb["Prezzi report sospetti"]):
        for col, w in larg.items():
            ws.column_dimensions[col].width = w
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                col = cell.column_letter
                if col in ("E",):
                    cell.number_format = "#,##0"
                elif col in ("F", "G", "H", "I", "J", "L", "M"):
                    cell.number_format = "#,##0.00"
                elif col == "N":
                    cell.number_format = "0.0"
        ws.freeze_panes = "A2"
    ws = wb["Dettaglio"]
    wb["Riepilogo"].column_dimensions["A"].width = 50
    wb["Riepilogo"].column_dimensions["B"].width = 18
    for c in ("C", "D", "E"):
        wb["Riepilogo"].column_dimensions[c].width = 16
    wb.save(OUT_XLSX)

    print(f"Scritto: {OUT_XLSX}")
    print(f"  Valore REPORT  : {tot_rep:,.2f}")
    print(f"  Nostro VALORE  : {tot_nos:,.2f}")
    print(f"  DELTA          : {tot_nos - tot_rep:,.2f}  ({(tot_nos-tot_rep)/tot_rep*100:+.2f}%)")
    print("\nPer fonte costo:")
    for _, r in g.iterrows():
        print(f"  {r['Fonte costo']:<16} n={int(r['n']):<5} report={r['valore_report']:>14,.2f} nostro={r['nostro_valore']:>14,.2f} delta={r['delta']:>12,.2f}")


if __name__ == "__main__":
    main()
